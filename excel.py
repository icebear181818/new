import openpyxl
import argparse
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("inputfile", help="input excel file for process")
    parser.add_argument("sheetname", help="corresponding sheet name for process")
    parser.add_argument("columnused", help="column need to be cleaned")
    parser.add_argument("max_row", help="total number of records")
    parser.add_argument("column", help="which column to save records")
    args = parser.parse_args()
    print args.inputfile
    print args.sheetname
    wb=openpyxl.load_workbook(args.inputfile)
    sheet=wb.get_sheet_by_name(args.sheetname)
    a=sheet.get_squared_range(int(args.columnused),2,int(args.columnused),int(args.max_row))
    vals = [[cell.value for cell in row] for row in a]
    tem1,tem2=[],[]
    for i in range(len(vals)):
        tem1.append(''.join(vals[i]).encode())

    for i in tem1:
        if "," in i:
            i = i.split(',')[0]
            tem2.append(i)
        elif i.endswith(('Corp.','limitation')):
            i = i.split()[0]
            tem2.append(i)
        else:
            tem2.append(i)
    for r in range(0,len(tem2)):
        sheet.cell(row=r+2,column=int(args.column)).value=tem2[r]
    wb.save(args.inputfile)
if __name__ == '__main__':
    main()

